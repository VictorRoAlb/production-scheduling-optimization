# ══════════════════════════════════════════════════════════════════════════════
# TRABAJO PPP — ARCHIVO 1: GENERACIÓN DE DATOS
# Máster UPV — Planificación y Programación de la Producción
# Grupo: Rafael Macian · Patricia Rodrigo · José Ramón Mena · Víctor Rodríguez
# ──────────────────────────────────────────────────────────────────────────────
# ► Ejecutar UNA SOLA VEZ con el Excel base.
# ► Lee los tiempos de proceso, genera todos los parámetros con semilla fija
#   y los escribe como nuevas hojas en el mismo Excel + un CSV por extensión.
# ══════════════════════════════════════════════════════════════════════════════

import os, csv, random      # os: rutas; csv: exportar; random: generación estocástica
import numpy as np          # media y operaciones vectoriales sobre los tiempos
import openpyxl             # lectura y escritura del archivo Excel


# ── CONFIGURACIÓN DE RUTAS ────────────────────────────────────────────────────
# Solo es necesario modificar RUTA_EXCEL; el resto se deriva automáticamente.

RUTA_EXCEL = os.environ.get("PPP_WORKBOOK_PATH", r".\\R_Cmax.xlsx")
# Ruta completa al archivo Excel que contiene la hoja base con los tiempos de proceso

DIR_CSV = os.path.join(os.path.dirname(RUTA_EXCEL), "datos_extensiones")
# Carpeta de salida para los CSVs — se crea en la misma ubicación que el Excel

# ─────────────────────────────────────────────────────────────────────────────


# Dimensiones del problema: 50 trabajos en 5 máquinas en paralelo no relacionadas
N = 50   # número de trabajos
M = 5    # número de máquinas

# Semillas fijas: garantizan que cualquier ejecución reproduce exactamente
# los mismos parámetros (criterio de reproducibilidad)
SEMILLA_RELEASES     = 123
SEMILLA_SETUPS       = 42
SEMILLA_DUEDATES     = 43
SEMILLA_PESOS        = 44
SEMILLA_PRECEDENCIAS = 45
SEMILLA_RECURSOS     = 46

os.makedirs(DIR_CSV, exist_ok=True)   # crea DIR_CSV si todavía no existe


# ══════════════════════════════════════════════════════════════════════════════
# 1. LEER TIEMPOS DE PROCESO DESDE EL EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def leer_tiempos_proceso():
    """
    Lee la hoja base del Excel.
    Devuelve p[i][j]: tiempo del trabajo j en la máquina i.
    Datos: fila 15 = primer trabajo, columna B = primera máquina.
    """
    wb   = openpyxl.load_workbook(RUTA_EXCEL, data_only=True)
    hoja = wb.active          # la hoja activa contiene la tabla de tiempos

    p = []
    for i in range(M):        # recorremos las M máquinas (columnas)
        fila_i = []
        for j in range(N):    # recorremos los N trabajos (filas desde la 15)
            valor = hoja.cell(row=15 + j, column=2 + i).value
            # Convertimos a entero: los tiempos de proceso son discretos;
            # leer como float podría introducir decimales espurios en el Cmax
            fila_i.append(int(round(float(valor))))
        p.append(fila_i)
    return p


# ══════════════════════════════════════════════════════════════════════════════
# 2. GENERADORES DE PARÁMETROS
# ══════════════════════════════════════════════════════════════════════════════

def generar_releases():
    """
    Divide los trabajos en tres grupos según cuándo están disponibles:
      40 % disponibles pronto   (r_j entre  0 y  5)
      35 % disponibles a mitad  (r_j entre  6 y 15)
      25 % disponibles tarde    (r_j entre 16 y 30)
    Después se mezclan aleatoriamente para que no queden agrupados.
    """
    random.seed(SEMILLA_RELEASES)
    n1, n2 = int(0.40 * N), int(0.35 * N)
    n3 = N - n1 - n2        # el resto cae en el bloque tardío

    # Concatenamos los tres grupos y los mezclamos aleatoriamente
    r = ( [random.randint( 0,  5) for _ in range(n1)] +
          [random.randint( 6, 15) for _ in range(n2)] +
          [random.randint(16, 30) for _ in range(n3)] )
    random.shuffle(r)
    return r


def generar_setups(p):
    """
    Genera tiempos de preparación aleatorios entre 1 y el 20 % de la media
    de los tiempos de proceso. La diagonal vale 0 porque preparar un trabajo
    tras sí mismo no tiene coste.
    """
    random.seed(SEMILLA_SETUPS)
    p_media = float(np.mean(p))
    s_max   = max(1, int(0.20 * p_media))   # límite superior = 20 % de la media

    setups = []
    for _ in range(M):
        # Matriz N×N: s[j][k] es el tiempo de preparación de j a k en esa máquina
        mat = [[0 if j == k else random.randint(1, s_max)
                for k in range(N)]
               for j in range(N)]
        setups.append(mat)
    return setups


def generar_due_dates(p, r):
    """
    La fecha de entrega de cada trabajo es su release date más un múltiplo
    aleatorio (entre 1.2 y 1.8) de su tiempo medio de proceso. Así todos
    los plazos son alcanzables pero con cierta presión de tiempo.
    """
    random.seed(SEMILLA_DUEDATES)
    d = []
    for j in range(N):
        p_media_j = float(np.mean([p[i][j] for i in range(M)]))
        alpha     = random.uniform(1.2, 1.8)   # holgura aleatoria por trabajo
        d.append(int(round(r[j] + alpha * p_media_j)))
    return d


def generar_pesos():
    """
    Asigna a cada trabajo un peso aleatorio entre 1 y 5 que representa
    su prioridad o importancia relativa en la función de tardanza ponderada.
    """
    random.seed(SEMILLA_PESOS)
    return [random.randint(1, 5) for _ in range(N)]


def generar_precedencias(n_pares=25):
    """
    Genera 25 restricciones de precedencia aleatorias sin ciclos.
    Para evitar ciclos solo se consideran pares donde j < k, de modo que
    el grafo es acíclico por construcción.
    """
    random.seed(SEMILLA_PRECEDENCIAS)
    # Al exigir j < k garantizamos que el grafo es acíclico por construcción
    candidatos = [(j, k) for j in range(N) for k in range(j + 1, N)]
    random.shuffle(candidatos)
    return candidatos[:n_pares]   # tomamos los primeros 25 tras mezclar


def generar_recursos():
    """
    Cada trabajo necesita entre 1 y 3 operarios para ejecutarse.
    La capacidad total H se fija al 60 % de la carga media por máquina,
    lo que crea una restricción real sin hacer el problema irresoluble.
    """
    random.seed(SEMILLA_RECURSOS)
    h = [random.randint(1, 3) for _ in range(N)]
    # H se fija al 60 % de la carga media por máquina para crear restricción real
    H = max(2, int(0.60 * sum(h) / M))
    return h, H


# ══════════════════════════════════════════════════════════════════════════════
# 3. ESCRIBIR EN EXCEL Y GUARDAR CSVs
# ══════════════════════════════════════════════════════════════════════════════

def _hoja_nueva(wb, nombre):
    # Si la hoja ya existe la eliminamos para sobreescribir sin duplicados
    if nombre in wb.sheetnames:
        del wb[nombre]
    return wb.create_sheet(nombre)


def guardar_csv(nombre_archivo, cabecera, filas):
    ruta = os.path.join(DIR_CSV, nombre_archivo)
    with open(ruta, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(cabecera)    # primera fila: nombres de columnas
        w.writerows(filas)      # resto: datos
    print(f"  CSV → {ruta}")


def escribir_todo(p, r, setups, d, w, prec, h, H):
    wb = openpyxl.load_workbook(RUTA_EXCEL)

    # ── Release dates ──────────────────────────────────────────────────────
    ws = _hoja_nueva(wb, "Release_dates")
    ws.cell(1, 1, "job"); ws.cell(1, 2, "rj")
    for j in range(N):
        ws.cell(j + 2, 1, j + 1)   # índice 1-based en Excel
        ws.cell(j + 2, 2, r[j])
    guardar_csv("release_dates.csv", ["job", "rj"],
                [[j + 1, r[j]] for j in range(N)])

    # ── Setups por máquina ─────────────────────────────────────────────────
    for i in range(M):
        ws = _hoja_nueva(wb, f"Setups_M{i + 1}")
        ws.cell(1, 1, f"Setups máquina {i + 1}  —  s[j_anterior][k_siguiente]")
        for j in range(N):
            for k in range(N):
                ws.cell(j + 2, k + 2, setups[i][j][k])
        guardar_csv(f"setups_M{i + 1}.csv",
                    [f"k{k}" for k in range(N)],
                    setups[i])

    # ── Due dates ──────────────────────────────────────────────────────────
    ws = _hoja_nueva(wb, "Due_dates")
    ws.cell(1, 1, "job"); ws.cell(1, 2, "dj")
    for j in range(N):
        ws.cell(j + 2, 1, j + 1)
        ws.cell(j + 2, 2, d[j])
    guardar_csv("due_dates.csv", ["job", "dj"],
                [[j + 1, d[j]] for j in range(N)])

    # ── Pesos ──────────────────────────────────────────────────────────────
    ws = _hoja_nueva(wb, "Pesos")
    ws.cell(1, 1, "job"); ws.cell(1, 2, "wj")
    for j in range(N):
        ws.cell(j + 2, 1, j + 1)
        ws.cell(j + 2, 2, w[j])
    guardar_csv("pesos.csv", ["job", "wj"],
                [[j + 1, w[j]] for j in range(N)])

    # ── Precedencias ───────────────────────────────────────────────────────
    ws = _hoja_nueva(wb, "Precedencias")
    ws.cell(1, 1, "j_antes"); ws.cell(1, 2, "k_despues")
    for idx, (j, k) in enumerate(prec):
        ws.cell(idx + 2, 1, j + 1)
        ws.cell(idx + 2, 2, k + 1)
    guardar_csv("precedencias.csv", ["j_antes", "k_despues"],
                [[j + 1, k + 1] for j, k in prec])

    # ── Recursos ───────────────────────────────────────────────────────────
    ws = _hoja_nueva(wb, "Recursos")
    ws.cell(1, 1, "job"); ws.cell(1, 2, "hj")
    ws.cell(1, 4, "H_total"); ws.cell(1, 5, H)   # H en celda aparte para leerlo fácil
    for j in range(N):
        ws.cell(j + 2, 1, j + 1)
        ws.cell(j + 2, 2, h[j])
    guardar_csv("recursos.csv", ["job", "hj", "H_total"],
                [[j + 1, h[j], H if j == 0 else ""] for j in range(N)])

    wb.save(RUTA_EXCEL)
    print(f"\n  Excel guardado → {RUTA_EXCEL}")


# ══════════════════════════════════════════════════════════════════════════════
# 4. MAIN
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("═" * 62)
    print("  GENERANDO PARÁMETROS")
    print("═" * 62)

    p    = leer_tiempos_proceso()
    r    = generar_releases()
    s    = generar_setups(p)
    d    = generar_due_dates(p, r)
    w    = generar_pesos()
    prec = generar_precedencias()
    h, H = generar_recursos()

    print("\nResumen:")
    print(f"  p_media                    = {np.mean(p):.2f}")
    print(f"  setup_max (M1, j!=k)       = {max(s[0][j][k] for j in range(N) for k in range(N) if j!=k)}")
    print(f"  r[:6]                      = {r[:6]}")
    print(f"  d[:6]                      = {d[:6]}")
    print(f"  w[:6]                      = {w[:6]}")
    print(f"  Nº precedencias            = {len(prec)}")
    print(f"  h[:6]                      = {h[:6]}")
    print(f"  H (capacidad personal)     = {H}")

    print("\nEscribiendo en Excel y CSVs...")
    escribir_todo(p, r, s, d, w, prec, h, H)

    print("\n✓ Generación completada. Ya puedes ejecutar 02_extensiones.py")
    print("═" * 62)



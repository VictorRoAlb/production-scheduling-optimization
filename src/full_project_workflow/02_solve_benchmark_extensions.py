 # ══════════════════════════════════════════════════════════════════════════════
# TRABAJO PPP — ARCHIVO 2: MÉTODOS Y EXTENSIONES
# Máster UPV — Planificación y Programación de la Producción
# Grupo: Rafael Macian · Patricia Rodrigo · José Ramón Mena · Víctor Rodríguez
# ──────────────────────────────────────────────────────────────────────────────
# ► Requiere haber ejecutado 01_generar_datos.py al menos una vez.
# ► Para ejecutar todo de golpe: python 02_extensiones.py
# ══════════════════════════════════════════════════════════════════════════════

import time, random, json, os   # time: medir ejecución; json/os: checkpoints
from collections import defaultdict

import numpy as np              # operaciones vectoriales auxiliares
import openpyxl                 # lectura del Excel con los datos del problema
from ortools.sat.python import cp_model   # solver CP-SAT de Google OR-Tools


# ── CONFIGURACIÓN DE RUTAS ────────────────────────────────────────────────────
# Solo es necesario modificar RUTA_EXCEL; el resto se deriva automáticamente.

RUTA_EXCEL = os.environ.get("PPP_WORKBOOK_PATH", r".\\R_Cmax.xlsx")

# Ruta completa al Excel con los datos del problema

RUTA_TXT        = os.path.join(os.path.dirname(RUTA_EXCEL), "resultados_ppp.txt")
# Fichero de salida con todos los resultados — misma carpeta que el Excel

RUTA_CHECKPOINT = os.path.join(os.path.dirname(RUTA_EXCEL), "checkpoint.json")
# Checkpoint de progreso — permite reanudar si se interrumpe la ejecución

# ─────────────────────────────────────────────────────────────────────────────


# Dimensiones del problema (deben coincidir con 01_generar_datos.py)
N = 50
M = 5

# ── Parámetros CP-SAT ─────────────────────────────────────────────────────────
TIEMPO_CPSAT  = 1000   # tiempo máximo de resolución en segundos
SEMILLA_CPSAT = 123    # semilla para reproducibilidad del solver
HILOS_CPSAT   = 1      # un solo hilo para garantizar determinismo
ESCALA_CPSAT  = 100    # factor de escala: convierte floats a enteros para CP-SAT

# ── Parámetros del Algoritmo Genético ─────────────────────────────────────────
SEMILLA_GEN   = 9      # semilla para reproducibilidad
TAM_POB       = 160    # tamaño de la población
P_CRUCE       = 0.92   # probabilidad de aplicar cruce OX
P_SWAP        = 0.30   # probabilidad de mutación por intercambio
P_INV         = 0.22   # probabilidad de mutación por inversión
ELITISMO      = 4      # individuos élite que pasan directamente a la siguiente generación
MAX_GEN       = 3000   # número máximo de generaciones
PACIENCIA_GEN = 350    # paradas sin mejora antes de terminar (criterio de parada)

# ── Parámetros del Iterated Greedy ───────────────────────────────────────────
SEMILLA_IG   = 42      # semilla para reproducibilidad
D_DEST       = 5       # trabajos eliminados en cada fase de destrucción
MAX_ITER_IG  = 800     # iteraciones máximas
PACIENCIA_IG = 200     # iteraciones sin mejora antes de terminar

print("✓ Configuración cargada.")


# ══════════════════════════════════════════════════════════════════════════════
# BLOQUE 1 — LECTURA DE DATOS
# ══════════════════════════════════════════════════════════════════════════════

def cargar_datos():
    wb   = openpyxl.load_workbook(RUTA_EXCEL, data_only=True)
    hoja = wb.active   # hoja base con la tabla de tiempos de proceso

    # p[i][j]: tiempo de proceso del trabajo j en la máquina i
    # Se convierte a entero para evitar Cmax decimal por artefactos de float en Excel
    p = [[int(round(float(hoja.cell(row=15 + j, column=2 + i).value)))
          for j in range(N)] for i in range(M)]

    ws = wb["Release_dates"]
    r  = [int(round(float(ws.cell(row=j + 2, column=2).value))) for j in range(N)]
    # r[j]: instante más pronto en que el trabajo j puede empezar a procesarse

    setups = []
    for i in range(M):
        ws  = wb[f"Setups_M{i + 1}"]
        # s[i][j][k]: tiempo de preparación al pasar del trabajo j al k en máquina i
        mat = [[int(round(float(ws.cell(row=j + 2, column=k + 2).value)))
                for k in range(N)] for j in range(N)]
        setups.append(mat)

    ws = wb["Due_dates"]
    d  = [int(round(float(ws.cell(row=j + 2, column=2).value))) for j in range(N)]
    # d[j]: fecha de entrega comprometida del trabajo j

    ws = wb["Pesos"]
    w  = [int(ws.cell(row=j + 2, column=2).value) for j in range(N)]
    # w[j]: peso o prioridad del trabajo j (relevante para la tardanza ponderada)

    ws   = wb["Precedencias"]
    prec = []
    fila = 2
    while ws.cell(row=fila, column=1).value is not None:
        j = int(ws.cell(row=fila, column=1).value) - 1   # pasamos a índice 0-based
        k = int(ws.cell(row=fila, column=2).value) - 1
        prec.append((j, k))
        fila += 1
    # prec: lista de pares (j, k) que indican que j debe terminar antes de que k empiece

    ws = wb["Recursos"]
    h  = [int(ws.cell(row=j + 2, column=2).value) for j in range(N)]
    H  = int(ws.cell(row=1, column=5).value)
    # h[j]: operarios requeridos por el trabajo j; H: capacidad total disponible

    return {'p': p, 'r': r, 'setups': setups,
            'd': d, 'w': w, 'precedencias': prec, 'h': h, 'H': H}


DATOS = cargar_datos()
print(f"✓ Datos cargados — {N} trabajos · {M} máquinas · H={DATOS['H']}")


# ══════════════════════════════════════════════════════════════════════════════
# BLOQUE 2 — FUNCIÓN CONSTRUCTIVA GREEDY
# ══════════════════════════════════════════════════════════════════════════════

def _esperar_recursos(t_prop, h_j, H, activos, duracion):
    """
    Retarda el inicio propuesto t_prop hasta que haya suficiente capacidad
    de personal disponible para ejecutar el trabajo (de duración 'duracion'
    y demanda h_j operarios) sin superar el límite H.
    """
    t = float(t_prop)
    for _ in range(1000):
        # Calculamos cuántos operarios están ocupados en el intervalo [t, t+dur)
        uso = sum(hk for (s, f, hk) in activos if s < t + duracion and f > t)
        if uso + h_j <= H:
            return t   # hay capacidad: inicio válido
        # Si no hay hueco, avanzamos al próximo evento de liberación
        futuros = [f for (_, f, _) in activos if f > t]
        if not futuros:
            break
        t = min(futuros)
    return t


def construir(perm, datos, cfg):
    """
    Función de evaluación greedy: dada una permutación de trabajos,
    los asigna uno a uno a la máquina que minimiza su tiempo de finalización,
    respetando las restricciones activas según cfg.
    Devuelve (objetivo, completions C, secuencias por máquina seq).
    """
    p      = datos['p']
    r      = datos['r']      if cfg.get('releases')  else [0.0] * N
    setups = datos['setups'] if cfg.get('setups')    else None
    h      = datos['h']      if cfg.get('recursos')  else None
    H      = datos['H']      if cfg.get('recursos')  else None

    # pred[b]: conjunto de trabajos que deben preceder a b
    pred = defaultdict(set)
    if cfg.get('precedencias'):
        for a, b in datos['precedencias']:
            pred[b].add(a)

    disponible = [0.0] * M   # instante en que cada máquina queda libre
    ultimo     = [-1]  * M   # último trabajo procesado en cada máquina (para setups)
    C          = [0.0] * N   # tiempo de finalización de cada trabajo
    seq        = defaultdict(list)
    activos    = []           # lista de (inicio, fin, operarios) para control de recursos

    for j in perm:
        # t_min: el trabajo no puede empezar antes que su release y sus predecesores
        t_min = float(r[j])
        if cfg.get('precedencias') and j in pred:
            for pre in pred[j]:
                t_min = max(t_min, C[pre])

        mejor_maq = None
        mejor_fin = float('inf')
        mejor_ini = 0.0

        for i in range(M):
            # Tiempo de setup: 0 si la máquina está vacía o no hay setups
            setup_ij = (setups[i][ultimo[i]][j]
                        if setups and ultimo[i] >= 0 else 0.0)
            ini = max(disponible[i] + setup_ij, t_min)
            if h is not None:
                ini = _esperar_recursos(ini, h[j], H, activos, p[i][j])
            fin = ini + p[i][j]

            # Desempate por menor índice
            if fin < mejor_fin - 1e-12 or (
                abs(fin - mejor_fin) <= 1e-12 and
                (mejor_maq is None or i < mejor_maq)
            ):
                mejor_fin = fin
                mejor_ini = ini
                mejor_maq = i

        C[j]                  = mejor_fin
        seq[mejor_maq].append(j)
        disponible[mejor_maq] = mejor_fin
        ultimo[mejor_maq]     = j
        if h is not None:
            activos.append((mejor_ini, mejor_fin, h[j]))

    # Función objetivo: Cmax o tardanza ponderada según cfg
    if cfg.get('tardanza'):
        d, w = datos['d'], datos['w']
        obj  = sum(w[j] * max(0.0, C[j] - d[j]) for j in range(N))
    else:
        obj = max(C)   # Cmax: tiempo de finalización del último trabajo

    return obj, C, dict(seq)


def reparar_prec(perm, prec):
    """
    Corrige una permutación para que respete todas las precedencias:
    si (a, b) ∈ prec y a aparece después de b, mueve a justo antes de b.
    """
    p   = list(perm)
    pos = {j: i for i, j in enumerate(p)}
    ok  = False
    while not ok:
        ok = True
        for a, b in prec:
            if a not in pos or b not in pos:
                continue
            if pos[a] > pos[b]:
                p.pop(pos[a])
                p.insert(pos[b], a)
                pos = {j: i for i, j in enumerate(p)}
                ok  = False   # reiniciamos el chequeo tras cada corrección
    return p


# ══════════════════════════════════════════════════════════════════════════════
# BLOQUE 3 — CP-SAT
# ══════════════════════════════════════════════════════════════════════════════

def resolver_cpsat(datos, cfg):
    """
    Formula el problema como un modelo CP-SAT de OR-Tools y lo resuelve
    dentro del límite de tiempo TIEMPO_CPSAT.
    Todas las variables se trabajan en unidades enteras escaladas (×ESCALA_CPSAT)
    para evitar la pérdida de precisión que tendría un modelo con floats.
    """
    p      = datos['p']
    r      = datos['r']      if cfg.get('releases')      else [0] * N
    setups = datos['setups'] if cfg.get('setups')        else None
    prec   = datos['precedencias'] if cfg.get('precedencias') else []
    h      = datos['h']      if cfg.get('recursos')      else None
    H      = datos['H']      if cfg.get('recursos')      else None
    d      = datos['d']      if cfg.get('tardanza')      else None
    w      = datos['w']      if cfg.get('tardanza')      else None

    # Escalar a enteros para CP-SAT (el solver solo trabaja con dominios enteros)
    ESC = ESCALA_CPSAT
    p_e = [[int(round(p[i][j] * ESC)) for j in range(N)] for i in range(M)]
    r_e = [int(round(r[j] * ESC)) for j in range(N)]
    s_e = ([[[int(round(setups[i][j][k] * ESC)) for k in range(N)]
              for j in range(N)] for i in range(M)] if setups else None)
    d_e = [int(round(d[j] * ESC)) for j in range(N)] if d else None

    # Horizonte temporal: cota superior holgada para los dominios de las variables
    max_p = max(p_e[i][j] for i in range(M) for j in range(N))
    max_r = max(r_e)
    if s_e:
        max_s = max(s_e[i][j][k]
                    for i in range(M) for j in range(N) for k in range(N))
        HT = max_r + N * (max_p + max_s) + 10
    else:
        HT = max_r + N * max_p + 10

    mdl  = cp_model.CpModel()

    # X[i,j] = 1 si el trabajo j se asigna a la máquina i
    X    = {(i, j): mdl.NewBoolVar(f"X_{i}_{j}")
            for i in range(M) for j in range(N)}
    S    = [mdl.NewIntVar(0, HT, f"S_{j}") for j in range(N)]   # inicio
    C    = [mdl.NewIntVar(0, HT, f"C_{j}") for j in range(N)]   # finalización
    Cmax = mdl.NewIntVar(0, HT, "Cmax")

    # Cada trabajo se asigna exactamente a una máquina
    for j in range(N):
        mdl.Add(sum(X[i, j] for i in range(M)) == 1)
    # Relación inicio–fin: C[j] = S[j] + p[maq][j]
    for j in range(N):
        mdl.Add(C[j] == S[j] + sum(p_e[i][j] * X[i, j] for i in range(M)))
    # Release dates: el trabajo no puede empezar antes de r[j]
    for j in range(N):
        mdl.Add(S[j] >= r_e[j])
    # Cmax >= C[j] para todo j
    for j in range(N):
        mdl.Add(Cmax >= C[j])

    # No solapamiento en máquinas con setups opcionales
    for i in range(M):
        for j in range(N):
            for k in range(j + 1, N):
                Y = mdl.NewBoolVar(f"Y_{i}_{j}_{k}")   # Y=1: j antes que k
                Z = mdl.NewBoolVar(f"Z_{i}_{j}_{k}")   # Z=1: ambos en máquina i
                mdl.Add(Z <= X[i, j])
                mdl.Add(Z <= X[i, k])
                mdl.Add(Z >= X[i, j] + X[i, k] - 1)
                sjk = s_e[i][j][k] if s_e else 0
                skj = s_e[i][k][j] if s_e else 0
                mdl.Add(S[k] >= C[j] + sjk - HT * (1 - Y) - HT * (1 - Z))
                mdl.Add(S[j] >= C[k] + skj - HT * Y       - HT * (1 - Z))

    # Restricciones de precedencia: b no puede empezar hasta que a termine
    for a, b in prec:
        mdl.Add(S[b] >= C[a])

    # Restricción de recursos: no superar H operarios simultáneos
    if h is not None:
        dur_vars = [mdl.NewIntVar(0, HT, f"dur_{j}") for j in range(N)]
        ivs      = []
        for j in range(N):
            mdl.Add(dur_vars[j] == sum(p_e[i][j] * X[i, j] for i in range(M)))
            ivs.append(mdl.NewIntervalVar(S[j], dur_vars[j], C[j], f"iv_{j}"))
        mdl.AddCumulative(ivs, h, H)

    # Función objetivo: tardanza ponderada o Cmax
    if cfg.get('tardanza'):
        T_vars  = [mdl.NewIntVar(0, HT, f"T_{j}") for j in range(N)]
        for j in range(N):
            mdl.AddMaxEquality(T_vars[j], [C[j] - d_e[j], mdl.NewConstant(0)])
        obj_var = mdl.NewIntVar(0, HT * N * 10, "obj")
        mdl.Add(obj_var == sum(w[j] * T_vars[j] for j in range(N)))
        mdl.Minimize(obj_var)
    else:
        mdl.Minimize(Cmax)

    slv = cp_model.CpSolver()
    slv.parameters.max_time_in_seconds = float(TIEMPO_CPSAT)
    slv.parameters.num_search_workers  = HILOS_CPSAT
    slv.parameters.random_seed         = SEMILLA_CPSAT

    t0     = time.time()
    estado = slv.Solve(mdl)
    t_tot  = round(time.time() - t0, 2)

    if estado not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return {'metodo': 'CP-SAT', 'objetivo': None, 'secuencia': {},
                'tiempo': t_tot, 'estado': slv.StatusName(estado)}

    raw   = slv.Value(obj_var) if cfg.get('tardanza') else slv.Value(Cmax)
    S_val = [slv.Value(S[j]) for j in range(N)]

    # Reconstruimos la secuencia por máquina ordenada por tiempo de inicio
    seq   = defaultdict(list)
    for j in range(N):
        maq = next(i for i in range(M) if slv.Value(X[i, j]) == 1)
        seq[maq].append(j)
    for i in range(M):
        seq[i].sort(key=lambda jj: S_val[jj])

    return {'metodo': 'CP-SAT', 'objetivo': raw // ESC,
            'tiempo': t_tot,    'estado': slv.StatusName(estado),
            'secuencia': dict(seq)}


# ══════════════════════════════════════════════════════════════════════════════
# BLOQUE 4 — ALGORITMO GENÉTICO
# ══════════════════════════════════════════════════════════════════════════════

def _cruce_ox(padre1, padre2):
    """
    Cruce de orden (OX): copia un segmento del padre1 y rellena el resto
    con los trabajos del padre2 en el orden en que aparecen.
    Preserva las posiciones relativas y garantiza permutaciones válidas.
    """
    n    = len(padre1)
    a, b = sorted(random.sample(range(n), 2))   # puntos de corte aleatorios
    hijo = [-1] * n
    hijo[a:b + 1] = padre1[a:b + 1]             # segmento heredado del padre1
    posicion = (b + 1) % n
    for trabajo in padre2:
        if trabajo not in hijo:
            hijo[posicion] = trabajo
            posicion = (posicion + 1) % n
    return hijo


def resolver_genetico(datos, cfg):
    """
    Algoritmo genético con selección por torneo, cruce OX, mutación por
    swap e inversión, y elitismo. La semilla fija garantiza reproducibilidad.
    """
    random.seed(SEMILLA_GEN)
    prec = datos['precedencias'] if cfg.get('precedencias') else []

    def evaluar(perm):
        # Si hay precedencias, reparamos la permutación antes de evaluarla
        p2 = reparar_prec(perm, prec) if prec else perm
        return construir(p2, datos, cfg)[0]

    # Permutación base ordenada por release date (heurística NEH simplificada)
    r    = datos['r']
    base = sorted(range(N), key=lambda j: (r[j], j)) if cfg.get('releases') \
           else sorted(range(N))

    # Inicialización mixta: 65 % cerca del óptimo greedy, 35 % aleatorio
    poblacion = []
    for _ in range(int(0.65 * TAM_POB)):
        perm = base[:]
        for _ in range(12):
            a, b = random.sample(range(N), 2)
            perm[a], perm[b] = perm[b], perm[a]   # pequeñas perturbaciones del greedy
        poblacion.append(perm)
    while len(poblacion) < TAM_POB:
        ind = list(range(N))
        random.shuffle(ind)
        poblacion.append(ind)

    fitness   = [evaluar(ind) for ind in poblacion]
    idx       = min(range(TAM_POB), key=lambda i: fitness[i])
    mejor_p   = poblacion[idx][:]
    mejor_f   = fitness[idx]
    gen_ult   = 0
    t0        = time.time()

    for gen in range(MAX_GEN):
        orden = sorted(range(TAM_POB), key=lambda i: fitness[i])
        nueva = [poblacion[i][:] for i in orden[:ELITISMO]]   # élite pasa intacta

        while len(nueva) < TAM_POB:
            def torneo():
                # Torneo de tamaño 3: el mejor de 3 individuos aleatorios
                c = random.sample(range(TAM_POB), 3)
                return poblacion[min(c, key=lambda i: fitness[i])][:]

            padre1, padre2 = torneo(), torneo()

            # ── 1 sola llamada random por cruce — igual que notebook ──
            if random.random() < P_CRUCE:
                hijo1 = _cruce_ox(padre1, padre2)
                hijo2 = _cruce_ox(padre2, padre1)
            else:
                hijo1 = padre1[:]
                hijo2 = padre2[:]

            # Mutación por swap: intercambia dos posiciones aleatorias
            if random.random() < P_SWAP:
                i, j = random.sample(range(N), 2)
                hijo1[i], hijo1[j] = hijo1[j], hijo1[i]
            if random.random() < P_SWAP:
                i, j = random.sample(range(N), 2)
                hijo2[i], hijo2[j] = hijo2[j], hijo2[i]

            # Mutación por inversión: invierte un subsegmento aleatorio
            if random.random() < P_INV:
                i, j = sorted(random.sample(range(N), 2))
                hijo1[i:j + 1] = hijo1[i:j + 1][::-1]
            if random.random() < P_INV:
                i, j = sorted(random.sample(range(N), 2))
                hijo2[i:j + 1] = hijo2[i:j + 1][::-1]

            nueva.append(hijo1)
            if len(nueva) < TAM_POB:
                nueva.append(hijo2)

        poblacion = nueva[:TAM_POB]
        fitness   = [evaluar(ind) for ind in poblacion]
        idx       = min(range(TAM_POB), key=lambda i: fitness[i])

        if fitness[idx] < mejor_f:
            mejor_f = fitness[idx]
            mejor_p = poblacion[idx][:]
            gen_ult = gen + 1

        # Criterio de parada por estancamiento
        if (gen + 1) - gen_ult >= PACIENCIA_GEN:
            break

    t_tot = round(time.time() - t0, 2)
    pf    = reparar_prec(mejor_p, prec) if prec else mejor_p
    _, _, seq = construir(pf, datos, cfg)

    return {'metodo': 'Genetico', 'objetivo': mejor_f,
            'tiempo': t_tot,      'estado': 'Metaheuristica',
            'secuencia': seq}


# ══════════════════════════════════════════════════════════════════════════════
# BLOQUE 5 — ITERATED GREEDY
# ══════════════════════════════════════════════════════════════════════════════

def resolver_ig(datos, cfg):
    """
    Iterated Greedy con búsqueda local por reinserción.
    Fase de destrucción: elimina D_DEST trabajos de la solución actual.
    Fase de construcción: los reinserta en la mejor posición greedy.
    Criterio de aceptación: solo mejora estricta.
    """
    random.seed(SEMILLA_IG)
    prec = datos['precedencias'] if cfg.get('precedencias') else []

    def evaluar(perm):
        p2 = reparar_prec(perm, prec) if prec else perm
        return construir(p2, datos, cfg)[0]

    def busqueda_local(perm):
        """
        Búsqueda local por reinserción: extrae cada trabajo y lo reinserta
        en la posición que minimiza el objetivo. Repite hasta que no mejore.
        """
        perm   = list(perm)
        obj    = evaluar(perm)
        mejoro = True
        while mejoro:
            mejoro = False
            for i in range(N):
                trab  = perm[i]
                resto = perm[:i] + perm[i + 1:]
                for pos in range(len(resto) + 1):
                    cand = resto[:pos] + [trab] + resto[pos:]
                    o2   = evaluar(cand)
                    if o2 < obj - 1e-9:
                        perm, obj = cand, o2
                        mejoro    = True
                        break
                if mejoro:
                    break
        return perm, obj

    # Solución inicial: order por release date si aplica, aleatorio si no
    if cfg.get('releases'):
        perm = sorted(range(N), key=lambda j: (datos['r'][j], j))
    else:
        perm = list(range(N))
        random.shuffle(perm)
    if prec:
        perm = reparar_prec(perm, prec)

    perm_best, obj_best = busqueda_local(perm)
    iter_sin_mej        = 0
    t0                  = time.time()

    for _ in range(MAX_ITER_IG):
        if iter_sin_mej >= PACIENCIA_IG:
            break

        # Destrucción: eliminamos D_DEST trabajos en posiciones aleatorias
        idx_el     = sorted(random.sample(range(N), D_DEST))
        eliminados = [perm_best[i] for i in idx_el]
        parcial    = [j for i, j in enumerate(perm_best)
                      if i not in set(idx_el)]

        # Construcción: reinsertamos los trabajos eliminados en el mejor hueco
        random.shuffle(eliminados)
        for trab in eliminados:
            mejor_pos = 0
            mejor_obj = float('inf')
            for pos in range(len(parcial) + 1):
                cand = parcial[:pos] + [trab] + parcial[pos:]
                if prec:
                    cand = reparar_prec(cand, prec)
                o2 = evaluar(cand)
                if o2 < mejor_obj - 1e-9:
                    mejor_obj = o2
                    mejor_pos = pos
            parcial = parcial[:mejor_pos] + [trab] + parcial[mejor_pos:]

        if prec:
            parcial = reparar_prec(parcial, prec)

        perm_nueva, obj_nueva = busqueda_local(parcial)

        if obj_nueva < obj_best - 1e-9:
            perm_best, obj_best = perm_nueva, obj_nueva
            iter_sin_mej        = 0
        else:
            iter_sin_mej += 1

    t_tot = round(time.time() - t0, 2)
    pf    = reparar_prec(perm_best, prec) if prec else perm_best
    _, _, seq = construir(pf, datos, cfg)

    return {'metodo': 'IG', 'objetivo': obj_best,
            'tiempo': t_tot, 'estado': 'Metaheuristica',
            'secuencia': seq}


# ══════════════════════════════════════════════════════════════════════════════
# BLOQUE 6 — RESULTADOS, RPDs, SECUENCIAS Y SALIDA TXT
# ══════════════════════════════════════════════════════════════════════════════

_txt_buffer = []   # acumula todas las líneas para escribirlas al final en el TXT


def _log(linea=""):
    """Imprime por pantalla y guarda en el buffer para el fichero TXT."""
    print(linea)
    _txt_buffer.append(linea)


def mostrar_resultados(resultados, titulo, label_obj="Cmax"):
    """
    Muestra tabla, RPDs y secuencias de TODOS los métodos de mejor a peor.
    RPD1 — 2º mejor vs mejor
    RPD2 — peor     vs mejor      (si hay 3 resultados)
    RPD3 — peor     vs 2º mejor   (si hay 3 resultados)
    """
    validos = sorted(
        [r for r in resultados if r['objetivo'] is not None],
        key=lambda r: r['objetivo']
    )

    SEP = "=" * 68
    _log()
    _log(SEP)
    _log(f"  {titulo}")
    _log(SEP)
    _log(f"  {'Metodo':<14} {'Estado':<26} {label_obj:>10} {'Tiempo (s)':>11}")
    _log(f"  {'-'*64}")
    for i, r in enumerate(validos):
        marca = "  << MEJOR" if i == 0 else ""
        _log(f"  {r['metodo']:<14} {r['estado']:<26} "
             f"{r['objetivo']:>10.2f} {r['tiempo']:>11.2f}{marca}")

    if len(validos) < 2:
        _log("  (Insuficientes resultados para RPDs)")
        _log(SEP)
        return

    mejor   = validos[0]
    segundo = validos[1]

    def rpd(cand, ref):
        # RPD = (candidato - referencia) / referencia × 100
        return (cand['objetivo'] - ref['objetivo']) / ref['objetivo'] * 100

    _log()
    _log(f"  -- RPDs " + "-" * 57)
    r1 = rpd(segundo, mejor)
    _log(f"  RPD1  {segundo['metodo']:>9}  vs  {mejor['metodo']:<10}"
         f"({segundo['objetivo']:.2f} - {mejor['objetivo']:.2f})"
         f" / {mejor['objetivo']:.2f} x 100 = {r1:+.2f}%")

    r2 = r3 = None
    if len(validos) == 3:
        peor = validos[2]
        r2   = rpd(peor, mejor)
        r3   = rpd(peor, segundo)
        _log(f"  RPD2  {peor['metodo']:>9}  vs  {mejor['metodo']:<10}"
             f"({peor['objetivo']:.2f} - {mejor['objetivo']:.2f})"
             f" / {mejor['objetivo']:.2f} x 100 = {r2:+.2f}%")
        _log(f"  RPD3  {peor['metodo']:>9}  vs  {segundo['metodo']:<10}"
             f"({peor['objetivo']:.2f} - {segundo['objetivo']:.2f})"
             f" / {segundo['objetivo']:.2f} x 100 = {r3:+.2f}%")

    # ── Secuencias de TODOS los métodos de mejor a peor ───────────────────
    _log()
    _log(f"  -- Secuencias por metodo (de mejor a peor) " + "-" * 22)
    for pos, r in enumerate(validos):
        etiqueta = ["MEJOR", "2o", "3o"][pos] if pos < 3 else f"{pos+1}o"
        _log(f"  [{etiqueta}] {r['metodo']}  ({label_obj}={r['objetivo']:.2f}):")
        seq = r.get('secuencia', {})
        for i in range(M):
            trabajos = [j + 1 for j in seq.get(i, [])]   # pasamos a índice 1-based
            _log(f"    M{i + 1}: {trabajos}")
        _log()

    _log(SEP)


def _cargar_ck():
    """Carga el checkpoint desde disco si existe; devuelve dict vacío si no."""
    if os.path.exists(RUTA_CHECKPOINT):
        with open(RUTA_CHECKPOINT, encoding="utf-8") as f:
            return json.load(f)
    return {}


def _guardar_ck(ck):
    """Persiste el checkpoint en disco tras completar cada extensión."""
    with open(RUTA_CHECKPOINT, "w", encoding="utf-8") as f:
        json.dump(ck, f, ensure_ascii=False, indent=2)


def _fix(r):
    """JSON serializa las claves de dict como strings; las restauramos a int."""
    if r.get('secuencia'):
        r['secuencia'] = {int(k): v for k, v in r['secuencia'].items()}
    return r


def guardar_txt():
    """
    Escribe el fichero TXT final con la cabecera de parámetros y todos
    los resultados acumulados en _txt_buffer.
    """
    cabecera = [
        "=" * 68,
        "  RESULTADOS COMPUTACIONALES -- TRABAJO PPP -- Master UPV",
        "  Generado por 02_extensiones.py",
        "=" * 68,
        "",
        f"  CP-SAT   semilla={SEMILLA_CPSAT}  hilos={HILOS_CPSAT}  "
        f"escala={ESCALA_CPSAT}  tiempo_max={TIEMPO_CPSAT}s",
        f"  Genetico semilla={SEMILLA_GEN}  tam_pob={TAM_POB}  "
        f"max_gen={MAX_GEN}  paciencia={PACIENCIA_GEN}",
        f"  IG       semilla={SEMILLA_IG}  D={D_DEST}  "
        f"max_iter={MAX_ITER_IG}  paciencia={PACIENCIA_IG}",
        "",
        "=" * 68,
        "",
    ]
    with open(RUTA_TXT, "w", encoding="utf-8") as f:
        f.write("\n".join(cabecera + _txt_buffer))
    print()
    print("=" * 68)
    print(f"  Resultados guardados en:")
    print(f"  {RUTA_TXT}")
    print("=" * 68)


# ══════════════════════════════════════════════════════════════════════════════
# EXTENSIONES
# ══════════════════════════════════════════════════════════════════════════════

# ── Checkpoint: carga progreso previo ─────────────────────────────────────────
_ck = _cargar_ck()
if _ck:
    print(f"  >> Checkpoint encontrado — extensiones ya hechas: "
          f"{[k for k in _ck]}")

# %% ── EXTENSIÓN 0: R || Cmax ─────────────────────────────────────────────────

cfg0 = dict(releases=False, setups=False,
            precedencias=False, recursos=False, tardanza=False)

if "ext0" in _ck:
    print("\n>>> EXTENSIÓN 0 ya calculada — cargando checkpoint...")
    r0_cp, r0_gen, r0_ig = [_fix(r) for r in _ck["ext0"]]
else:
    print("\n>>> Ejecutando EXTENSION 0: R || Cmax ...")
    r0_cp  = resolver_cpsat    (DATOS, cfg0)
    r0_gen = resolver_genetico (DATOS, cfg0)
    r0_ig  = resolver_ig       (DATOS, cfg0)
    _ck["ext0"] = [r0_cp, r0_gen, r0_ig]
    _guardar_ck(_ck)
mostrar_resultados([r0_cp, r0_gen, r0_ig],
                   "EXTENSION 0 -- R || Cmax", "Cmax")


# %% ── EXTENSIÓN 1: R | rj | Cmax ────────────────────────────────────────────

cfg1 = dict(releases=True,  setups=False,
            precedencias=False, recursos=False, tardanza=False)

if "ext1" in _ck:
    print("\n>>> EXTENSIÓN 1 ya calculada — cargando checkpoint...")
    r1_cp, r1_gen, r1_ig = [_fix(r) for r in _ck["ext1"]]
else:
    print("\n>>> Ejecutando EXTENSION 1: R | rj | Cmax ...")
    r1_cp  = resolver_cpsat    (DATOS, cfg1)
    r1_gen = resolver_genetico (DATOS, cfg1)
    r1_ig  = resolver_ig       (DATOS, cfg1)
    _ck["ext1"] = [r1_cp, r1_gen, r1_ig]
    _guardar_ck(_ck)
mostrar_resultados([r1_cp, r1_gen, r1_ig],
                   "EXTENSION 1 -- R | rj | Cmax", "Cmax")


# %% ── EXTENSIÓN 2: R | rj, sijk | Cmax ─────────────────────────────────────

cfg2 = dict(releases=True,  setups=True,
            precedencias=False, recursos=False, tardanza=False)

if "ext2" in _ck:
    print("\n>>> EXTENSIÓN 2 ya calculada — cargando checkpoint...")
    r2_cp, r2_gen, r2_ig = [_fix(r) for r in _ck["ext2"]]
else:
    print("\n>>> Ejecutando EXTENSION 2: R | rj, sijk | Cmax ...")
    r2_cp  = resolver_cpsat    (DATOS, cfg2)
    r2_gen = resolver_genetico (DATOS, cfg2)
    r2_ig  = resolver_ig       (DATOS, cfg2)
    _ck["ext2"] = [r2_cp, r2_gen, r2_ig]
    _guardar_ck(_ck)
mostrar_resultados([r2_cp, r2_gen, r2_ig],
                   "EXTENSION 2 -- R | rj, sijk | Cmax", "Cmax")


# %% ── EXTENSIÓN 3: R | rj, sijk, prec | Cmax ───────────────────────────────

cfg3 = dict(releases=True,  setups=True,
            precedencias=True, recursos=False, tardanza=False)

if "ext3" in _ck:
    print("\n>>> EXTENSIÓN 3 ya calculada — cargando checkpoint...")
    r3_cp, r3_gen, r3_ig = [_fix(r) for r in _ck["ext3"]]
else:
    print("\n>>> Ejecutando EXTENSION 3: R | rj, sijk, prec | Cmax ...")
    r3_cp  = resolver_cpsat    (DATOS, cfg3)
    r3_gen = resolver_genetico (DATOS, cfg3)
    r3_ig  = resolver_ig       (DATOS, cfg3)
    _ck["ext3"] = [r3_cp, r3_gen, r3_ig]
    _guardar_ck(_ck)
mostrar_resultados([r3_cp, r3_gen, r3_ig],
                   "EXTENSION 3 -- R | rj, sijk, prec | Cmax", "Cmax")


# %% ── EXTENSIÓN 4: R | rj, sijk, prec, res | Cmax ──────────────────────────

cfg4 = dict(releases=True,  setups=True,
            precedencias=True, recursos=True, tardanza=False)

if "ext4" in _ck:
    print("\n>>> EXTENSIÓN 4 ya calculada — cargando checkpoint...")
    r4_cp, r4_gen, r4_ig = [_fix(r) for r in _ck["ext4"]]
else:
    print("\n>>> Ejecutando EXTENSION 4: R | rj, sijk, prec, res | Cmax ...")
    r4_cp  = resolver_cpsat    (DATOS, cfg4)
    r4_gen = resolver_genetico (DATOS, cfg4)
    r4_ig  = resolver_ig       (DATOS, cfg4)
    _ck["ext4"] = [r4_cp, r4_gen, r4_ig]
    _guardar_ck(_ck)
mostrar_resultados([r4_cp, r4_gen, r4_ig],
                   "EXTENSION 4 -- R | rj, sijk, prec, res | Cmax", "Cmax")


# %% ── EXTENSIÓN 5: R | rj, sijk, prec, res | sum(wjTj) ─────────────────────

cfg5 = dict(releases=True,  setups=True,
            precedencias=True, recursos=True, tardanza=True)

if "ext5" in _ck:
    print("\n>>> EXTENSIÓN 5 ya calculada — cargando checkpoint...")
    r5_cp, r5_gen, r5_ig = [_fix(r) for r in _ck["ext5"]]
else:
    print("\n>>> Ejecutando EXTENSION 5: R | rj, sijk, prec, res, d_j | sum(wjTj) ...")
    r5_cp  = resolver_cpsat    (DATOS, cfg5)
    r5_gen = resolver_genetico (DATOS, cfg5)
    r5_ig  = resolver_ig       (DATOS, cfg5)
    _ck["ext5"] = [r5_cp, r5_gen, r5_ig]
    _guardar_ck(_ck)
mostrar_resultados([r5_cp, r5_gen, r5_ig],
                   "EXTENSION 5 -- R | rj, sijk, prec, res,d_j | sum(wjTj)", "sum(wjTj)")


# ── GUARDAR TXT ───────────────────────────────────────────────────────────────
guardar_txt()


